"""
Run once to add the Email column to the sample transactions file.

Usage (from project root):
    python scripts/add_email_column.py
"""
import shutil
from pathlib import Path

import openpyxl

SRC = Path.home() / "Downloads" / "Copy of CCCoordinator_transactionFile.xlsx"
DSTS = [
    Path("data") / "transactions_sample.xlsx",
    Path("tests") / "fixtures" / "sample_transactions.xlsx",
]

for DST in DSTS:
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)

    wb = openpyxl.load_workbook(DST)
    ws = wb["Main Report"]

    # Add Email header in column H
    ws.cell(row=1, column=8, value="Email")

    # Placeholder addresses — replace at least one with your real email
    EMAIL_MAP = {
        "CORY BACH":      "cory.bach@placeholder.com",
        "TODD BAHR":      "todd.bahr@placeholder.com",
        "ARIEL BENEWIAT": "ariel.benewiat@placeholder.com",
    }
    for row in ws.iter_rows(min_row=2):
        name = row[0].value
        if name in EMAIL_MAP:
            ws.cell(row=row[0].row, column=8, value=EMAIL_MAP[name])

    wb.save(DST)
    print(f"Saved → {DST}")

print("\nDone. Open data/transactions_sample.xlsx and replace a placeholder")
print("email with your real address so you can test the end-to-end flow.")
